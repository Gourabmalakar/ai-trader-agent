from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Time (IST)", "time"),
    ("Symbol", "symbol"),
    ("Name", "name"),
    ("Side", "side"),
    ("Quantity", "quantity"),
    ("Price", "price"),
    ("Cost basis", "costBasis"),
    ("Realized P&L", "realizedPnl"),
    ("Status", "status"),
    ("Engine", "provider"),
    ("Reason", "reason"),
]


def build_trade_log_xlsx(trades: list[dict[str, Any]]) -> Response:
    """Render the trade log (with reasoning) as a real .xlsx workbook, returned as a downloadable
    FastAPI Response — not a CSV-pretending-to-be-Excel, an actual spreadsheet file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trade log"

    for col_index, (header, _key) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)

    for row_index, trade in enumerate(trades, start=2):
        for col_index, (_header, key) in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=col_index, value=trade.get(key))

    for col_index, (header, key) in enumerate(COLUMNS, start=1):
        width = max(len(header) + 2, 14 if key != "reason" else 60)
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ai-trader-agent-trade-log.xlsx"},
    )
